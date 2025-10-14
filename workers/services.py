from typing import Any, Dict

from django.db.models import QuerySet
from django.db import IntegrityError, transaction

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Worker
from .serializers import WorkerImportRowSerializer


class WorkerService:
    @staticmethod
    def get_workers() -> QuerySet[Worker]:
        """Получить всех работников."""
        try:
            with transaction.atomic():
                workers = Worker.objects.all()
                return workers
        except IntegrityError as exc:
            raise exc

    @staticmethod
    def get_worker(pk) -> Worker:
        """Получить работника по ID."""
        try:
            with transaction.atomic():
                workers = Worker.objects.get(id=pk)
                return workers
        except IntegrityError as exc:
            raise exc
        
    @staticmethod
    def import_workers_from_excel(file_obj, created_by) -> Dict[str, Any]:
        """Импорт работников из Excel файла."""
        workbook = load_workbook(filename=file_obj, read_only=True, data_only=True)
        sheet = workbook.active

        headers: List[str] = []
        first_row = True
        required_fields = {'first_name', 'last_name', 'email', 'position'}

        added_count = 0
        total_rows = 0
        errors: List[Dict[str, Any]] = []

        existing_emails = set(Worker.objects.values_list('email', flat=True))
        seen_emails = set()

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if first_row:
                headers = [str(h).strip() if h is not None else '' for h in row]
                missing = required_fields - set(headers)
                if missing:
                    return {
                        'added': 0,
                        'errors': [{
                            'detail': f'Отсутствуют обязательные столбцы: {", ".join(sorted(missing))}'
                        }]}
                first_row = False
                continue
            total_rows += 1
            row_data: Dict[str, Any] = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    row_data[headers[col_idx]] = value

            serializer = WorkerImportRowSerializer(data=row_data)
            if not serializer.is_valid():
                errors.append({'row': row_idx, 'detail': serializer.errors})
                continue

            email_value = str(serializer.validated_data.get('email')).strip().lower()
            if email_value in seen_emails:
                errors.append({'row': row_idx, 'detail': 'Дубликат email в файле'})
                continue
            if email_value in existing_emails:
                errors.append({'row': row_idx, 'detail': 'Email уже существует в базе'})
                continue

            try:
                with transaction.atomic():
                    worker = Worker(**serializer.to_worker_kwargs())
                    worker.created_by = created_by
                    worker.save()
                    added_count += 1
                    seen_emails.add(email_value)
            except Exception as exc:
                errors.append({'row': row_idx, 'detail': str(exc)})

        return {
            'added': added_count,
            'errors': errors,
            'total': total_rows,
        }